import io
from openpyxl import load_workbook
from .base import BaseProcessor, ExtractionResult


class XlsxProcessor(BaseProcessor):
    async def extract(self, file_content: bytes, filename: str) -> ExtractionResult:
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        all_text = []
        tables = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cell_values = [
                    str(cell) if cell is not None else "" for cell in row
                ]
                if any(v.strip() for v in cell_values):
                    rows.append(cell_values)

            if rows:
                all_text.append(f"[Sheet: {sheet_name}]")
                headers = rows[0]
                all_text.append(" | ".join(headers))
                for row in rows[1:]:
                    all_text.append(" | ".join(row))

                tables.append({
                    "sheet": sheet_name,
                    "headers": headers,
                    "rows": rows[1:],
                })

        return ExtractionResult(
            text="\n".join(all_text),
            tables=tables,
            metadata={
                "filename": filename,
                "sheet_count": len(wb.sheetnames),
            },
        )
